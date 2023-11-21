from .extractor import *
import openpyxl

class ExcelCellExtractor(Extractor):
    method = 'excelCell'
    def __init__(self):
        pass
    def extract(self, path: str) -> ExtractResult:
        search_texts = []
        extract_details = {
        }
        file_texts = []

        all_sheet_data = extract_excel_cell(path)
        if all_sheet_data is None:
            return None
        for i, (sheet_name, sheet) in enumerate(all_sheet_data.items()):
            page_texts = []
            for cell in sheet:
                word_details = {
                    "sheet_name": sheet_name,
                    "address": cell["address"],
                    "row": cell["row"],
                    "column": cell["column"],
                    "sheetCount": len(all_sheet_data.items()),
                    "sheetNumber": i,
                }
                word_text = str(cell["value"])
                word_search_text = SearchText(word_text, SearchTextUnit.word, word_details)
                search_texts.append(word_search_text)
                page_texts.append(word_text)
                file_texts.append(word_text)
            page_details = {
                "sheetName": sheet_name,
                "sheetCount": len(all_sheet_data.items()),
            }
            page_search_text = SearchText('\n'.join(page_texts), SearchTextUnit.page, page_details)
            search_texts.append(page_search_text)
        file_details = {
            "sheetCount": len(all_sheet_data.items()),
        }
        file_search_text = SearchText('\n'.join(file_texts), SearchTextUnit.file, file_details)
        search_texts.append(file_search_text)
        
        return ExtractResult(extract_details, search_texts)

def extract_cell_data(cell) -> dict[str, any]:
    if cell.data_type not in ["s", "n", "inlineStr", "str"] or cell.value is None or cell.value == "":
        return None
    return {
        "address": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "value": cell.value,
        # "formula": cell.formula if cell.data_type == 'f' else None
    }

def extract_excel_cell(file_path) -> dict:
  # Excelファイルを開く
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(e)
        return None

    # セルのデータを格納するためのリスト
    all_sheet_data = {}

    # すべてのシートに対して処理を行う
    for i, sheet_name in enumerate(list(workbook.sheetnames)):
        sheet = workbook[sheet_name]
        all_cell_data = []

        # シート内のセルに対して処理を行う
        for row in sheet.iter_rows():
            for cell in row:
                cell_data = extract_cell_data(cell)
                if cell_data is not None:
                  all_cell_data.append(cell_data)
        all_sheet_data[sheet_name] = all_cell_data

    return all_sheet_data